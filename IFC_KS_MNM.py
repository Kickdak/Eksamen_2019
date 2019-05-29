import PySimpleGUI as sg
import sys
import ifcopenshell
import os
from openpyxl import load_workbook
import uuid
import time


txtfil = "filbane.txt"
with open('readme.txt','r') as readme:
    brk_veil =readme.read()
def filbane():
        window_rows = [[sg.Text('Velg IFC Lokasjon')],
                       [sg.InputText(), sg.FileBrowse()],
                       [sg.Submit("Kjør"), sg.Cancel("Exit")]]
        window = sg.Window("IFC_KS_M'N'M")
        # Loop for programmet og knapper
        while True:  # Event Loop<
            event, values = window.Layout(window_rows).Read()
            if event in (None, 'Exit'):
                sys.exit("aa! errors!")
            if event == 'Kjør':
                window.Close()
                break
        global source_filename
        source_filename = values[0]
        with open(txtfil,"w") as writefile:
            writefile.write(source_filename+"\n")

if not os.path.exists(txtfil):
    filbane()


def pset_creater():
    txtfil = "filbane.txt"
    global parameter_fil
    parameter_fil = "KS_punkter.xlsx"

    wb = load_workbook(parameter_fil)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    fagfelt = []
    klar = ['Avvik',]
    index = 2
    for f in range(2, max_column + 1):
        fag = sheet.cell(row=1, column=f)
        fagfelt.append(fag.value)
        for i in range(index, max_row + 1):
            cell_obj = sheet.cell(row=i, column=int(f))
            if not cell_obj.value == None:
                sjekk = cell_obj.value
                klar.append(sjekk)
    print(klar)

    def filbane():

        window_rows = [[sg.Text('Velg IFC Lokasjon')],
                       [sg.InputText(), sg.FileBrowse()],
                       [sg.Submit("Kjør"), sg.Cancel("Exit")]]
        window = sg.Window("IFC_KS_M'N'M")

        # Loop for programmet og knapper
        while True:  # Event Loop<
            event, values = window.Layout(window_rows).Read()
            if event in (None, 'Exit'):
                sys.exit("aa! errors!")
            if event == 'Kjør':
                window.Close()
                break
        global source_filename
        source_filename = values[0]
        with open(txtfil, "w") as writefile:
            writefile.write(source_filename + "\n")

    if not os.path.exists(txtfil):
        filbane()

    with open(txtfil, "r") as txt:
        global fil
        fil = txt.read().replace("\n", "")
    ifcfile = ifcopenshell.open(fil)
    create_guid = lambda: ifcopenshell.guid.compress(uuid.uuid1().hex)
    owner_history = ifcfile.by_type("IfcOwnerHistory")[0]
    spaces = ifcfile.by_type("IfcSpace")
    sets = ifcfile.by_type("IfcPropertySet")
    list = []
    for sett in sets:
        list.append(sett.Name)
    myset = set(list)


    ########### Properties Lages  ###############
    def create_pset():
        file = parameter_fil
        global antall_parametere
        antall_parametere = 0
        global index
        index = 0
        antall_parametere = len(klar)
        print(antall_parametere)

        if "Pset_KS" not in myset:
            print("Pset lages")

            property_values = []

            print(property_values)
            antall_properties = len(property_values)
            print(antall_properties)
            for space in spaces:
                while index < antall_parametere:
                    property_values.append(ifcfile.createIfcPropertySingleValue(klar[index], klar[index],ifcfile.create_entity("IfcText","Ikke klar"),None), )

                    index += 1
                property_set = ifcfile.createIfcPropertySet(create_guid(), owner_history, "Pset_KS", "Pset_KS",property_values)
                ifcfile.createIfcRelDefinesByProperties(create_guid(), owner_history, None, None, [space], property_set)

                ifcfile.write(fil)
                property_values = []
                index = 0

        else:
            print("Pset finnes")
            sg.PopupError("Pset er allerede oprettet i modell.")

    create_pset()
    ifcOrganizer()
##### Brukerveiledning #####
def brukerveil():
    layout3 = [[sg.Text(brk_veil)],[sg.Cancel('Lukk')]]

    window3 = sg.Window('Brukerveiledning',keep_on_top=True,location=(575,173),no_titlebar=True).Layout(layout3)

    while True:
        event, values = window3.Read()
        if event == 'Lukk':
            window3.Close()
            break


##### Avviks håndtering #####
def avvik(rom, fagfelt, parametere):

    def reg_avvik(rom, fagfelt, parametere, avvik_txt):
        dir = os.getcwd()
        dato = time.strftime(" %Y-%m-%d Kl %H.%M", time.localtime())
        avviks_navn = rom +' - '+ fagfelt + '.txt'
        avvik_skille = parametere +' - '+ dato+':'
        avviksmappe = dir+'/avviks_meldinger'
        global avvik_logg
        global logg_mappe
        avvik_logg = rom+' - Avvikslogg.txt'
        logg_mappe = dir+'/avviks_logg'

        ######## HVIS AVVIKSMAPPE IKKE ALLEREDE FINNES BLIR DETTE LAGET HER ########
        if not os.path.exists(avviksmappe):
            os.makedirs(avviksmappe)

        ######## AVVIK DELT INN I FAGFELT BLIR REGISTRERT HER OG LAGT INN I AVVIKSMAPPE ########
        with open(avviksmappe+'/'+avviks_navn,'a') as add_txt:
            add_txt.write(avvik_skille+'\n'+avvik_txt+'\n')
            print ('Notert - Fag_avvik')

        ######## HVIS LOGG MAPPE IKKE ALLEREDE FINNES BLIR DETTE LAGET HER ########
        if not os.path.exists(logg_mappe):
            os.makedirs(logg_mappe)
        ######## AVVIK DELT INN PR ROM BLIR REGISTRERT HER OG LAGT INN I LOGG MAPPE ########
        with open(logg_mappe + '/' + avvik_logg, 'a') as add_txt:
            add_txt.write(fagfelt+' - '+avvik_skille + '\n' + avvik_txt + '\n')
            print('Notert - logg')

    layout = [[sg.Text('Avviks Registrering:')],[sg.Multiline('', size=(96,10),key='avvik_txt')],[sg.Submit('Registrer'),sg.Cancel('Avbryt')]]

    window2 = sg.Window('Avviks melding',keep_on_top=True,no_titlebar=True,grab_anywhere=False,location=(575, 173)).Layout(layout)
    global IfcRelSorted

    while True:
        event,values = window2.Read()
        if event == 'Avbryt':
            window2.Close()
            break
        if event == 'Registrer':
            reg_avvik(rom,fagfelt,parametere,values['avvik_txt'])
            print ('Avvik registrert')
            oppdater_parameter('Avvik', rom, logg_mappe + '/' + avvik_logg, ifcfile,IfcRelSorted , fil)
            window2.Close()
            break


def ifcOrganizer():
    global fil
    with open(txtfil,"r") as txt:
        fil = txt.read().replace("\n","")
    global ifcfile
    ifcfile = ifcopenshell.open(fil)
    spaces = ifcfile.by_type("IfcSpace")
    global sets
    sets = ifcfile.by_type("IfcPropertySet")
    liste = []
    romliste = []
    propertyliste = []


    ###HENTER ROM VERDIER OG PROPERTIES TIL LISTER###
    for room in spaces:
        romliste.append(room.Name)
    for sett in sets:
        liste.append(sett.Name)
    global IfcRelSorted
    for sett in sets:
        if sett.Name == "Pset_KS":
            for job in sett.__getitem__(4):
                propertyliste.append(job.__getitem__(0))
    IfcRelUnsorted = ifcfile.by_type("IfcRelDefinesByProperties")
    IfcRelSorted = []
    for test in IfcRelUnsorted:
        if test.__getitem__(5).__getitem__(2) == "Pset_KS":
            IfcRelSorted.append(test)

    ###SORTERING AV MYLIST, MY LIST ER = PROPERTYLISTE###
    mylist = list(dict.fromkeys(propertyliste))
    mylist.sort()
    print(mylist)
    ###STARTER CREATER HVIS IKKE PSET FINNES###
    if mylist == []:
        print("hey")
        sg.Popup('Pset_KS ble ikke i funnet.\nOppretter Pset_KS og parametere.',auto_close=True,auto_close_duration=5,button_type=5)
        pset_creater()
        sys.exit()

    ##### EXCEL INNHENTING AV DATA ####
    filepath = "KS_punkter.xlsx"
    wb = load_workbook(filepath)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    ### GLOBAL BRUKES FÅR Å BRUKE VARIABLER I ANDRE FUNKSJONER ###
    global fagfelt
    global klar
    global klar
    global sjekkliste
    global sjekk
    fagfelt = []
    klar = []
    sjekkliste = []
    index = 2  # Rad
    ####### SØKER IGJENNOM EXCEL ARK ETTER INFROMASJON #######
    for f in range(2, max_column + 1):
        # print (f)
        fag = sheet.cell(row=1, column=f)
        fagfelt.append(fag.value)
    ########### LAGES SOM EN PLACEHOLDER FOR FØRSTE VERDI I PARAMETER DROPDOWN ##########
    for i in range(index,max_row+1):
        cell_obj=sheet.cell(row=i,column=2)
        if not cell_obj.value == None:
            sjekk = cell_obj.value
            klar.append(sjekk)

    global layout

    nåverdi=False
    #################### UI STARTER ER ####################
    layout = [
        [sg.Text('Rom nr: ', size=(6, 1)), sg.Combo(romliste,key="rom"),sg.Text('Fagfelt: ', size=(5, 1)), sg.Combo(fagfelt, change_submits=True,key="fagfelt"), sg.Text('Punkt: '),
         sg.Combo(klar, key='parametere'), sg.Checkbox('Utført',key="Utført"), sg.Checkbox('Avvik',key="Avvik"),sg.Submit(" OK "), sg.Cancel("Exit"),sg.RButton("",key="Help",image_filename="info_icon.png",button_color=sg.TRANSPARENT_BUTTON,image_subsample=8, border_width=0),sg.RButton("",key="Pin",image_filename="pin.png",button_color=sg.TRANSPARENT_BUTTON,image_subsample=20, border_width=0)],
    ]

    buildingElement = sg.Window("IFC_KS_M'N'M", keep_on_top=True, resizable=True, location=(575, 130),no_titlebar=True, grab_anywhere=False).Refresh()

    # Loop for programmet og knapper
    while True:  # Event Loop<
        event, values = buildingElement.Layout(layout).Read()
        if event in (None, 'Exit'):
            print("Skriptet ble stoppet")
            sys.exit()
        if event == "Help":
            brukerveil()
        if event == "Pin":
            if nåverdi == False:
                buildingElement.GrabAnyWhereOn()
                nåverdi = True
            elif nåverdi == True:
                buildingElement.GrabAnyWhereOff()
                nåverdi = False
        if event == " OK ":
            if values["Utført"] == False and values["Avvik"] == True:
                values["ferdig"] = "Ikke klart med Avvik"
                avvik(values['rom'], values['fagfelt'],values['parametere'])
            elif values["Utført"] == True and values["Avvik"] == True:
                values["ferdig"] = "Ferdig med Avvik"
                avvik(values['rom'], values['fagfelt'],values['parametere'])
            elif values["Utført"] == False:
                values["ferdig"] = "Ikke klart"
            elif values["Utført"] == True:
                values["ferdig"] = "Ferdig"
            oppdater_parameter(values["parametere"], values["rom"], values["ferdig"], ifcfile, IfcRelSorted, fil)

        nedtrekk = values["fagfelt"]
        fagnr = fagfelt.index(nedtrekk)
        klar = []
        for i in range(index, max_row + 1):
            cell_obj = sheet.cell(row=i, column=int(fagnr + 2))
            if not cell_obj.value == None:
                sjekk = cell_obj.value
                klar.append(sjekk)
        buildingElement.FindElement('parametere').Update(values=klar)


def oppdater_parameter(parametere, romnummer, ferdig, ifc, sets, fil):

    for sett in sets:
        if sett.__getitem__(4).__getitem__(0).__getitem__(2) == romnummer:
            print(sett)
            for job in sett.__getitem__(5).__getitem__(4):
                if job.Name == parametere:
                    print(job)
                    job.NominalValue.__setitem__(0, ferdig)
                    print(job.NominalValue)

    print("skille")
    ifc.write(fil)
if __name__ == '__main__':
    ifcOrganizer()

#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
#####################################       LAGET AV HENRIK NYSTRØM, THOMAS MAGRAFF OG PEDER MUGAAS       #####################################
